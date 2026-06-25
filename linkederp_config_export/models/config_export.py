# -*- coding: utf-8 -*-
import io
import re
import base64
import zipfile
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
# ─── Palette ─────────────────────────────────────────────────────────────────
C = {
    "hdr_bg": "1F3864", "hdr_fg": "FFFFFF",
    "sub_bg": "2E75B6", "sub_fg": "FFFFFF",
    "sect":   "D6E4F0", "alt":    "EBF5FB",
    "done_bg": "C6EFCE", "done_fg": "375623",
    "prog_bg": "FFEB9C", "prog_fg": "9C5700",
    "ns_bg":  "FFC7CE", "ns_fg":  "9C0006",
    "na_bg":  "F2F2F2", "na_fg":  "595959",
    "white":  "FFFFFF", "bdr":    "BDC3C7",
    "sa_grn": "007A4D",
}
HEADERS = [
    "Module / Area", "Configuration Item", "Menu Path / Location",
    "Setting / Action Required", "Current Value (from DB)",
    "Best Practice / Recommendation", "Data Migration Required",
    "Owner", "Status", "Notes",
]
COL_W = [22, 35, 42, 48, 38, 48, 14, 16, 15, 32]
STATUS_DV = '"Not Started,In Progress,Done,N/A,Blocked"'
MIGR_DV   = '"Yes,No,Partial,TBD"'
OWNER_DV  = '"AL3 Team,Consultant,IT,Finance,HR,Operations,Management,Payroll"'
def _fill(h):
    return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=11):
    return Font(bold=bold, color=color, size=size, name="Calibri")
def _align(h="left", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=True)
def _border():
    s = Side(style="thin", color=C["bdr"])
    return Border(left=s, right=s, top=s, bottom=s)
def _status_fill(status):
    m = {
        "Done":        (C["done_bg"], C["done_fg"]),
        "In Progress": (C["prog_bg"], C["prog_fg"]),
        "Not Started": (C["ns_bg"],   C["ns_fg"]),
        "N/A":         (C["na_bg"],   C["na_fg"]),
        "Blocked":     ("FF0000",     "FFFFFF"),
    }
    return m.get(status, (C["white"], "000000"))
def _dv(formula, col_letter, max_row=500):
    d = DataValidation(type="list", formula1=formula, allow_blank=True)
    d.sqref = f"{col_letter}3:{col_letter}{max_row}"
    return d
class Al3ConfigExport(models.TransientModel):
    _name = "al3.config.export"
    _description = "AL3 Configuration Export Wizard"

    name = fields.Char(
        string="Export Name",
        default=lambda self: "AL3 Config " + datetime.now().strftime("%Y-%m-%d"),
    )
    area_ids = fields.Many2many(
        "al3.config.area",
        string="Configuration Areas",
        help="Areas are auto-selected based on installed modules. You can adjust manually.",
    )
    company_id = fields.Many2one(
        "res.company", string="Company",
        default=lambda self: self.env.company,
    )
    include_cover = fields.Boolean("Include Cover Sheet", default=True)
    state = fields.Selection([("draft", "Draft"), ("done", "Done")], default="draft")
    # ── Excel output ──────────────────────────────────────────────────────────
    excel_file = fields.Binary("Excel File", readonly=True)
    file_name  = fields.Char("File Name", readonly=True)
    # ── Module output ─────────────────────────────────────────────────────────
    module_file     = fields.Binary("Module (.zip)", readonly=True)
    module_filename = fields.Char("Module Filename", readonly=True)
    module_name     = fields.Char("Generated Module Name", readonly=True)
    # ── Export options ─────────────────────────────────────────────────────
    export_partners = fields.Boolean(
        "Export Partners (Customers/Vendors)", default=False,
        help="Include customer and vendor records in the module export (max 200 each).",
    )
    export_products = fields.Boolean(
        "Export Products", default=False,
        help="Include product templates in the module export (max 500).",
    )
    installed_modules_info = fields.Char(
        string="Installed Modules",
        compute="_compute_installed_modules_info",
    )

    @api.depends("area_ids")
    def _compute_installed_modules_info(self):
        installed = self.env["ir.module.module"].search(
            [("state", "=", "installed"), ("application", "=", True)]
        )
        names = ", ".join(installed.mapped("shortdesc")[:10])
        suffix = f" (+{len(installed) - 10} more)" if len(installed) > 10 else ""
        for rec in self:
            rec.installed_modules_info = f"{len(installed)} apps: {names}{suffix}"

    # ── Menu-driven area detection ────────────────────────────────────────────
    # Areas come from Odoo's own menu tree: one area per root app (top-level
    # menu) that exposes a "Configuration" branch. Nothing is hard-coded per
    # module — each area's content is read live from the models its menus open.
    _AREA_COLORS = [
        "1F3864", "2E75B6", "375623", "9C5700", "007A4D", "C00000",
        "7030A0", "0070C0", "00B0F0", "FFC000", "ED7D31", "595959",
        "833C00", "002060", "244185", "538135", "843C0C", "4472C4",
    ]

    def _menu_config_branch(self, root_menu):
        """Return the 'Configuration' child of a root menu, if any."""
        for child in root_menu.child_id:
            if (child.name or "").strip().lower() == "configuration":
                return child
        return False

    def _get_auto_detected_area_ids(self):
        """Build one area per root app that has a Configuration menu branch."""
        ConfigArea = self.env["al3.config.area"].sudo()
        roots = self.env["ir.ui.menu"].search(
            [("parent_id", "=", False)], order="sequence, name"
        )
        selected, live_codes = [], set()
        for idx, root in enumerate(roots):
            if not self._menu_config_branch(root):
                continue
            code = f"menu_{root.id}"
            live_codes.add(code)
            vals = {
                "name": (root.name or code)[:60],
                "code": code,
                "sequence": idx,
                "description": f"Live configuration under the {root.name} app",
                "tab_color": self._AREA_COLORS[idx % len(self._AREA_COLORS)],
                "active": True,
            }
            area = ConfigArea.search([("code", "=", code)], limit=1)
            if area:
                area.write(vals)
            else:
                area = ConfigArea.create(vals)
            selected.append(area.id)
        # Deactivate any area that no longer maps to a current root menu
        for area in ConfigArea.search([("active", "=", True)]):
            if area.code not in live_codes:
                area.active = False
        return selected

    # ── Override default_get → auto-populate areas on wizard open ────────────
    @api.model
    def default_get(self, fields_list):
        res = super().default_get(fields_list)
        if "area_ids" in fields_list:
            res["area_ids"] = [(6, 0, self._get_auto_detected_area_ids())]
        return res

    # ── Buttons ───────────────────────────────────────────────────────────────
    def action_select_all(self):
        self.area_ids = self.env["al3.config.area"].search([])

    def action_clear_all(self):
        self.area_ids = [(5, 0, 0)]

    def action_auto_detect(self):
        """Re-run auto-detection and refresh the form."""
        self.area_ids = [(6, 0, self._get_auto_detected_area_ids())]
        return {
            "type": "ir.actions.act_window",
            "res_model": "al3.config.export",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
    def action_export(self):
        if not HAS_OPENPYXL:
            raise UserError(_(
                "The openpyxl library is required.\n"
                "Please install it on the server:\n"
                "  pip install openpyxl"
            ))
        if not self.area_ids:
            raise UserError(_("Please select at least one configuration area."))
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        if self.include_cover:
            self._build_cover(wb)
        for area in self.area_ids.sorted("sequence"):
            rows = self._extract_area(area)
            if rows:
                self._build_sheet(wb, area, rows)
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        company_safe = (self.company_id.name or "AL3").replace(" ", "_")
        filename = company_safe + '_OdooConfig_' + datetime.now().strftime('%Y%m%d_%H%M') + '.xlsx'
        self.excel_file = base64.b64encode(buf.read())
        self.file_name  = filename
        self.state      = "done"
        return {
            "type": "ir.actions.act_window",
            "res_model": "al3.config.export",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }
    def action_download(self):
        return {
            "type": "ir.actions.act_url",
            "url": (
                f"/web/content/al3.config.export/{self.id}"
                f"/excel_file/{self.file_name}?download=true"
            ),
            "target": "self",
        }

    def action_download_module(self):
        return {
            "type": "ir.actions.act_url",
            "url": (
                f"/web/content/al3.config.export/{self.id}"
                f"/module_file/{self.module_filename}?download=true"
            ),
            "target": "self",
        }

    # ══════════════════════════════════════════════════════════════════════════
    # MODULE GENERATOR  –  generates an installable Odoo module (.zip)
    # ══════════════════════════════════════════════════════════════════════════
    def action_generate_module(self):
        """
        Generate a ready-to-install Odoo module containing:
          - __manifest__.py
          - __init__.py
          - data/*.xml  (one file per configuration area with real DB values)
        The module can be installed on any Odoo instance / Odoo Online.
        """
        if not self.area_ids:
            raise UserError(_("Please select at least one configuration area first."))

        comp = self.company_id
        company_slug = re.sub(r"[^a-z0-9]+", "_",
                               (comp.name or "project").lower()).strip("_")
        mod_name     = f"config_{company_slug}_exported"
        dt_str       = datetime.now().strftime("%Y%m%d_%H%M")

        buf = io.BytesIO()
        data_files = []
        depends     = {"base"}

        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:

            def add(path, content):
                zf.writestr(f"{mod_name}/{path}", content)

            # ── 1. Company settings ───────────────────────────────────────────
            xml = self._modgen_company()
            if xml:
                add("data/01_company.xml", xml)
                data_files.append("data/01_company.xml")

            # ── 2. Master data per app (menu-driven, same source as Excel) ─────
            dumped_models = set()
            for seq, area in enumerate(self.area_ids.sorted("sequence"), start=10):
                root = self._root_menu_of_area(area)
                if not root:
                    continue
                xml = self._modgen_area(area, root, dumped_models, depends)
                if xml:
                    slug = re.sub(r"[^a-z0-9]+", "_", (area.name or "app").lower()).strip("_")
                    path = f"data/{seq:02d}_{slug}.xml"
                    add(path, xml)
                    data_files.append(path)

            # ── 3. ir.config_parameter records ───────────────────────────────
            xml = self._modgen_settings_params()
            if xml:
                add("data/90_settings.xml", xml)
                data_files.append("data/90_settings.xml")

            # ── 4. res.partner (customers, suppliers) ────────────────────────
            if self.export_partners:
                xml = self._modgen_partners()
                if xml:
                    add("data/05_partners.xml", xml)
                    data_files.append("data/05_partners.xml")

            # ── 5. Products ──────────────────────────────────────────────────
            if self.export_products:
                xml = self._modgen_products()
                if xml:
                    add("data/06_products.xml", xml)
                    data_files.append("data/06_products.xml")

            # ── 6. post_init_hook: res.config.settings.execute() ─────────────
            # This is the only reliable way to apply settings (implied_group,
            # module toggles, compute side-effects) on the target database.
            settings_vals = self._capture_settings_values()
            hook_py = self._modgen_settings_hook(settings_vals)
            if hook_py:
                add("hooks.py", hook_py)
                has_hook = True
            else:
                has_hook = False

            # ── 7. __init__.py — import hook so Odoo can find it ─────────────
            init_content = "# Auto-generated by AL3 Config Export\n"
            if has_hook:
                init_content += "from . import hooks\n"
            add("__init__.py", init_content)

            # ── 8. __manifest__.py ────────────────────────────────────────────
            installed_apps = self.env["ir.module.module"].search(
                [("state", "=", "installed"), ("application", "=", True)]
            )
            # depends = ALL installed modules so Project B auto-installs
            # everything before loading XML data and running the hook.
            # Without this, XML refs to account/tax/uom records fail on
            # a fresh Project B that has none of those modules installed.
            all_installed = self.env["ir.module.module"].search(
                [("state", "=", "installed")]
            ).mapped("name")
            # Exclude our own exporter module and technical base modules
            # that are always present (no need to declare them).
            _always_present = {"base", "web", "mail", "bus", "digest",
                               "linkederp_config_export", mod_name}
            full_depends = sorted(
                m for m in all_installed if m not in _always_present
            )
            data_files_sorted = sorted(set(data_files))
            manifest = self._modgen_manifest(
                mod_name, comp, full_depends,
                data_files_sorted, installed_apps,
                post_init_hook="post_init_hook" if has_hook else None,
            )
            add("__manifest__.py", manifest)

        buf.seek(0)
        filename = f"{mod_name}_{dt_str}.zip"
        self.module_file     = base64.b64encode(buf.read())
        self.module_filename = filename
        self.module_name     = mod_name
        self.state           = "done"
        return {
            "type": "ir.actions.act_window",
            "res_model": "al3.config.export",
            "res_id": self.id,
            "view_mode": "form",
            "target": "new",
        }

    # ── XML helpers ───────────────────────────────────────────────────────────
    @staticmethod
    def _xe(val):
        """Escape a value for XML."""
        if val is None:
            return ""
        return (str(val)
                .replace("&", "&amp;")
                .replace("<", "&lt;")
                .replace(">", "&gt;")
                .replace('"', "&quot;"))

    @staticmethod
    def _xml_wrap(body, noupdate=False):
        # ponytail: noupdate=False so records overwrite existing data in the target DB
        nu = "1" if noupdate else "0"
        return (
            '<?xml version="1.0" encoding="utf-8"?>\n'
            "<odoo>\n"
            f'    <data noupdate="{nu}">\n'
            f"{body}\n"
            "    </data>\n"
            "</odoo>\n"
        )

    @staticmethod
    def _xml_record(xml_id, model, fields_xml, indent=8):
        pad = " " * indent
        p2  = " " * (indent - 4)
        return (
            f'{p2}<record id="{xml_id}" model="{model}">\n'
            f"{fields_xml}\n"
            f"{p2}</record>"
        )

    def _xf(self, name, value, indent=12):
        """One <field> line."""
        pad = " " * indent
        return f'{pad}<field name="{name}">{self._xe(value)}</field>'

    def _xf_ref(self, name, ref, indent=12):
        pad = " " * indent
        return f'{pad}<field name="{name}" ref="{ref}"/>'

    # ── __manifest__.py generator ─────────────────────────────────────────────
    def _modgen_manifest(self, mod_name, comp, depends, data_files, installed_apps,
                         post_init_hook=None):
        ordered_deps = ["base"] + [d for d in depends if d != "base"]
        dep_list  = ",\n        ".join(f'"{d}"' for d in ordered_deps)
        data_list = ",\n        ".join(f'"{f}"' for f in data_files)
        app_names = ", ".join(installed_apps.mapped("shortdesc")[:10])
        hook_line = f'\n    "post_init_hook": "{post_init_hook}",' if post_init_hook else ""
        today = datetime.now().strftime("%Y-%m-%d")
        now   = datetime.now().strftime("%Y-%m-%d %H:%M")
        return (
            f'# -*- coding: utf-8 -*-\n'
            f'# Auto-generated by AL3 Config Export on {now}\n'
            f'# Source: {self._xe(comp.name)} | Generated from: al3_config_export\n'
            f'{{\n'
            f'    "name": "Config: {self._xe(comp.name)} (Exported)",\n'
            f'    "version": "19.0.1.0.0",\n'
            f'    "category": "Technical/Configuration",\n'
            f'    "summary": "Configuration snapshot exported from {self._xe(comp.name)}",\n'
            f'    "description": (\n'
            f'        "Auto-generated configuration module. "\n'
            f'        "Applies settings, master data, and configuration "\n'
            f'        "captured from {self._xe(comp.name)} on {today}. "\n'
            f'        "Installed apps in source: {self._xe(app_names)}."\n'
            f'    ),\n'
            f'    "author": "AL3 Config Export",\n'
            f'    "depends": [\n'
            f'        {dep_list}\n'
            f'    ],\n'
            f'    "data": [\n'
            f'        {data_list}\n'
            f'    ],{hook_line}\n'
            f'    "installable": True,\n'
            f'    "application": False,\n'
            f'    "auto_install": False,\n'
            f'    "license": "LGPL-3",\n'
            f'}}\n'
        )

    # ── Company XML ───────────────────────────────────────────────────────────
    def _modgen_company(self):
        c = self.company_id
        fields_xml = "\n".join(filter(None, [
            self._xf("name",    c.name)     if c.name     else "",
            self._xf("vat",     c.vat)      if c.vat      else "",
            self._xf("phone",   c.phone)    if c.phone    else "",
            self._xf("email",   c.email)    if c.email    else "",
            self._xf("website", c.website)  if c.website  else "",
            self._xf("street",  c.street)   if c.street   else "",
            self._xf("city",    c.city)     if c.city     else "",
            self._xf("zip",     c.zip)      if c.zip      else "",
            (self._xf_ref("country_id", self._ext_id_for(c.country_id))
             if c.country_id and self._ext_id_for(c.country_id) else ""),
            (self._xf_ref("currency_id", self._ext_id_for(c.currency_id))
             if c.currency_id and self._ext_id_for(c.currency_id) else ""),
        ]))
        if not fields_xml.strip():
            return None
        rec = self._xml_record("base.main_company", "res.company", fields_xml)
        comment = "    <!-- ── Company Settings ── -->"
        return self._xml_wrap(f"{comment}\n{rec}")

    # ── res.config.settings live values ──────────────────────────────────────
    def _capture_settings_values(self):
        """
        Return a dict of {fname: value} for all applicable res.config.settings
        fields, read from the live database. Used by both XML and hook generators.
        Only includes fields that have a non-default value worth exporting.
        """
        Settings = self.env["res.config.settings"]
        try:
            cfg = Settings.sudo().with_context(
                active_id=self.company_id.id
            ).create({"company_id": self.company_id.id})
        except Exception:
            return {}
        result = {}
        skip = {"id", "display_name", "company_id", "company_ids", "__last_update"}
        for fname, field_obj in sorted(Settings._fields.items()):
            if fname in skip or fname.startswith("_"):
                continue
            if not field_obj.store and not getattr(field_obj, "config_parameter", None) \
                    and not getattr(field_obj, "implied_group", None):
                # Skip purely computed/non-persistent fields that carry no config
                ftype = field_obj.type
                if ftype not in ("boolean", "selection", "char", "integer", "float", "many2one"):
                    continue
            try:
                val = cfg[fname]
            except Exception:
                continue
            ftype = field_obj.type
            if ftype == "boolean":
                result[fname] = bool(val)
            elif ftype in ("char", "text", "selection"):
                if val not in (False, None, ""):
                    result[fname] = val
            elif ftype in ("integer", "float"):
                if val not in (False, None, 0, 0.0):
                    result[fname] = val
            elif ftype == "many2one":
                if val:
                    ext = self._ext_id_for(val)
                    if ext:
                        result[fname] = ("ref", ext)
        return result

    def _modgen_settings_params(self):
        """
        Two outputs in one file:
        1. ir.config_parameter records for fields with config_parameter key
           (the XML-loadable subset — safe to write as data records).
        2. Returns the settings dict so action_generate_module can build the hook.
        This method only writes the ir.config_parameter XML part.
        """
        try:
            Settings = self.env["res.config.settings"]
            lines = ["    <!-- ── ir.config_parameter values ── -->"]
            count = 0
            for fname, field_obj in sorted(Settings._fields.items()):
                param_key = getattr(field_obj, "config_parameter", None)
                if not param_key:
                    continue
                param = self.env["ir.config_parameter"].sudo().get_param(param_key)
                if param in (False, None, ""):
                    continue
                xml_id = re.sub(r"[^a-z0-9]", "_", param_key)
                f_xml = "\n".join([
                    self._xf("key",   param_key),
                    self._xf("value", str(param)),
                ])
                lines.append(self._xml_record(
                    f"cfg_param_{xml_id}", "ir.config_parameter", f_xml
                ))
                count += 1
            if count == 0:
                return None
            return self._xml_wrap("\n".join(lines), noupdate=False)
        except Exception:
            return None

    def _modgen_settings_hook(self, settings_vals):
        """
        Generate a post_init_hook Python function that calls
        res.config.settings.create({...}).execute() on install.
        This is the ONLY reliable way to activate Odoo settings that rely on
        implied_group, module install/uninstall, or compute side-effects.
        """
        if not settings_vals:
            return None
        lines = [
            "# -*- coding: utf-8 -*-",
            "# Auto-generated by AL3 Config Export — applies res.config.settings",
            "",
            "",
            "def post_init_hook(env):",
            '    """Apply configuration settings captured from the source database."""',
            "    vals = {",
        ]
        for fname, val in sorted(settings_vals.items()):
            if isinstance(val, tuple) and val[0] == "ref":
                safe_ref = val[1].replace("'", "\\'")
                # resolve ref at install time; None entries are filtered below
                lines.append(
                    f"        '{fname}': (_r := env.ref('{safe_ref}', raise_if_not_found=False)) and _r.id,"
                )
            elif isinstance(val, bool):
                lines.append(f"        '{fname}': {val},")
            elif isinstance(val, str):
                safe = val.replace("\\", "\\\\").replace("'", "\\'")
                lines.append(f"        '{fname}': '{safe}',")
            else:
                lines.append(f"        '{fname}': {val},")
        lines += [
            "    }",
            "    # Remove None values (unresolved refs)",
            "    vals = {k: v for k, v in vals.items() if v is not None}",
            "    try:",
            "        env['res.config.settings'].sudo().create(vals).execute()",
            "    except Exception as e:",
            "        import logging",
            "        logging.getLogger(__name__).warning('Config hook failed: %s', e)",
        ]
        return "\n".join(lines) + "\n"

    # ── Accounting XML ────────────────────────────────────────────────────────
    # ── Generic, menu-driven master-data dumper ───────────────────────────────
    # These mirror the Excel extractor: walk an app's Configuration menu, and
    # for every model it opens, emit installable XML records. To guarantee the
    # generated module installs cleanly we are conservative:
    #   • only stored, writable, non-computed fields are exported;
    #   • relations (m2o/m2m) are emitted ONLY when the target already has an
    #     external id (ref=) OR is part of this same export — otherwise skipped;
    #   • a record whose *required* relation cannot be resolved is skipped whole.

    # Field names never worth writing (technical / auto-managed).
    _DUMP_SKIP_FIELDS = {
        "create_uid", "write_uid", "create_date", "write_date", "__last_update",
        "display_name", "access_token", "message_main_attachment_id",
    }
    # Models we never try to (re)create as data — too entangled / auto-built,
    # transactional, or child rows that only exist under a parent record.
    _DUMP_SKIP_RELATION_MODELS = {
        "mail.alias", "mail.activity", "mail.message", "mail.followers",
        "ir.actions.report", "ir.actions.act_window", "ir.ui.view",
        "stock.rule", "stock.picking.type", "account.online.account",
        "calendar.event", "product.supplierinfo", "link.tracker",
        "res.partner", "res.users", "product.template", "product.product",
    }

    def _ext_id_for(self, record):
        """Return the *real* 'module.name' external id for a record, or None.

        Odoo's transient ``__export__`` ids are ignored — they are not stable
        and can't be depended on by the generated module.
        """
        if not record:
            return None
        imd = self.env["ir.model.data"].sudo().search(
            [("model", "=", record._name), ("res_id", "=", record.id),
             ("module", "!=", "__export__")], limit=1
        )
        return f"{imd.module}.{imd.name}" if imd else None

    def _local_xmlid(self, record):
        """Stable in-module xml id for a record we are creating ourselves."""
        model_slug = record._name.replace(".", "_")
        return f"{model_slug}_{record.id}"

    def _resolve_ref(self, rec, idmap):
        """External id usable in XML for a related record, or None.

        A relation resolves if the target already has a real external id, or if
        we are creating it ourselves in this export (it's in ``idmap``).
        """
        if not rec:
            return None
        return self._ext_id_for(rec) or idmap.get((rec._name, rec.id))

    def _record_resolvable(self, record, idmap):
        """True if every *required* relation of the record can be referenced."""
        for fname, f in self.env[record._name]._fields.items():
            if f.type not in ("many2one", "many2many"):
                continue
            if not getattr(f, "required", False):
                continue
            if not f.store or f.compute is not None or f.related:
                continue
            if f.comodel_name in self._DUMP_SKIP_RELATION_MODELS:
                return False
            val = record[fname]
            if f.type == "many2one":
                if val and not self._resolve_ref(val, idmap):
                    return False
            else:  # required m2m → need at least one resolvable target
                if val and not any(self._resolve_ref(r, idmap) for r in val):
                    return False
        return True

    def _dump_fields(self, record, idmap):
        """Build the <field> lines for one record (assumed resolvable)."""
        lines = []
        for fname, f in sorted(self.env[record._name]._fields.items()):
            if fname in self._DUMP_SKIP_FIELDS:
                continue
            if not f.store or f.compute is not None or f.related:
                continue
            if getattr(f, "automatic", False):
                continue
            try:
                val = record[fname]
            except Exception:
                continue

            if f.type == "many2one":
                if f.comodel_name in self._DUMP_SKIP_RELATION_MODELS or not val:
                    continue
                ref = self._resolve_ref(val, idmap)
                if ref:
                    lines.append(self._xf_ref(fname, ref))
            elif f.type == "many2many":
                if f.comodel_name in self._DUMP_SKIP_RELATION_MODELS or not val:
                    continue
                refs = [self._resolve_ref(r, idmap) for r in val]
                refs = [r for r in refs if r]
                if refs:
                    joined = ", ".join(f"ref('{r}')" for r in refs)
                    lines.append(
                        f'            <field name="{fname}" '
                        f'eval="[(6, 0, [{joined}])]"/>'
                    )
            elif f.type == "one2many":
                continue  # children handled when their own model is dumped
            elif f.type == "boolean":
                lines.append(self._xf(fname, "1" if val else "0"))
            elif f.type in ("char", "text", "html", "selection"):
                if val in (False, None, ""):
                    continue
                lines.append(self._xf(fname, val))
            elif f.type in ("integer", "float", "monetary"):
                if val in (False, None):
                    continue
                lines.append(self._xf(fname, val))
            elif f.type == "date":
                if val:
                    lines.append(self._xf(fname, val))
            # datetime/binary/properties → skipped (rarely config, often noisy)
        return "\n".join(lines)

    def _modgen_area(self, area, root, dumped_models, depends):
        """Emit one XML file for an app, dumping its config models' records."""
        cfg = self._menu_config_branch(root)
        if not cfg:
            return None
        module = self._module_of_menu(root)
        if module:
            depends.add(module)

        # ── Collect candidate records (only those we'd actually create) ────────
        # Records that already have a real external id are referenced, never
        # recreated — so seeded/core data (countries, base tags…) can't collide.
        records_by_model = []          # [(leaf_name, model, [records to create])]
        for full_path, leaf_name, model in self._iter_config_leaves(cfg, root.name):
            if model in self._SKIP_MODELS or model in dumped_models:
                continue
            if model not in self.env or model in self._DUMP_SKIP_RELATION_MODELS:
                continue
            dumped_models.add(model)
            try:
                Model = self.env[model].sudo()
                recs = Model.search(self._company_domain(model),
                                    limit=self._RECORD_CAP)
            except Exception:
                continue
            # Include ALL records: those with ext IDs become updates, new ones become creates
            if recs:
                records_by_model.append((leaf_name, model, list(recs)))

        # ── Plan xml ids; ext-id records use their real id, new records get local one ──
        idmap = {}
        ext_id_set = set()
        for _, m, recs in records_by_model:
            for r in recs:
                ext = self._ext_id_for(r)
                idmap[(m, r.id)] = ext if ext else self._local_xmlid(r)
                if ext:
                    ext_id_set.add((m, r.id))
        # Drop only NEW records with unresolvable required relations
        changed = True
        while changed:
            changed = False
            for _, model, recs in records_by_model:
                for r in list(recs):
                    key = (model, r.id)
                    if key in idmap and key not in ext_id_set and not self._record_resolvable(r, idmap):
                        del idmap[key]
                        recs.remove(r)
                        changed = True

        # ── Emit survivors ─────────────────────────────────────────────────────
        lines = [f"    <!-- ── {area.name} ── -->"]
        wrote = 0
        for leaf_name, model, recs in records_by_model:
            if not recs:
                continue
            seg = [f"    <!-- {leaf_name} ({model}) -->"]
            seg_wrote = 0
            for r in recs:
                fields_xml = self._dump_fields(r, idmap)
                if not fields_xml.strip():
                    continue
                seg.append(self._xml_record(idmap[(model, r.id)], model, fields_xml))
                seg_wrote += 1
            if seg_wrote:
                lines.extend(seg)
                wrote += seg_wrote

        if not wrote:
            return None
        return self._xml_wrap("\n".join(lines))

    # ── Partners (customers, vendors, contacts) XML ─────────────────────────
    def _modgen_partners(self):
        cid   = self.company_id.id
        lines = []
        try:
            # Customers
            customers = self.env["res.partner"].search(
                [("customer_rank", ">", 0)],
                limit=200, order="name"
            )
            if customers:
                lines.append("    <!-- ── Partners: Customers ── -->")
                for i, p in enumerate(customers):
                    xml_id = f"partner_customer_{i:04d}"
                    f_xml  = "\n".join(filter(None, [
                        self._xf("name", p.name) if p.name else "",
                        self._xf("email", p.email) if p.email else "",
                        self._xf("phone", p.phone) if p.phone else "",
                        self._xf("street", p.street) if p.street else "",
                        self._xf("city", p.city) if p.city else "",
                        self._xf("zip", p.zip) if p.zip else "",
                        (self._xf_ref("country_id", self._ext_id_for(p.country_id))
                         if p.country_id and self._ext_id_for(p.country_id) else ""),
                        (self._xf_ref("state_id", self._ext_id_for(p.state_id))
                         if p.state_id and self._ext_id_for(p.state_id) else ""),
                        self._xf("vat", p.vat) if p.vat else "",
                        self._xf("website", p.website) if p.website else "",
                    ]))
                    if f_xml.strip():
                        lines.append(self._xml_record(xml_id, "res.partner", f_xml))
        except Exception:
            pass
        try:
            # Vendors
            vendors = self.env["res.partner"].search(
                [("supplier_rank", ">", 0)],
                limit=200, order="name"
            )
            if vendors:
                lines.append("    <!-- ── Partners: Vendors ── -->")
                for i, p in enumerate(vendors):
                    xml_id = f"partner_vendor_{i:04d}"
                    f_xml  = "\n".join(filter(None, [
                        self._xf("name", p.name) if p.name else "",
                        self._xf("email", p.email) if p.email else "",
                        self._xf("phone", p.phone) if p.phone else "",
                        self._xf("street", p.street) if p.street else "",
                        self._xf("city", p.city) if p.city else "",
                        self._xf("zip", p.zip) if p.zip else "",
                        (self._xf_ref("country_id", self._ext_id_for(p.country_id))
                         if p.country_id and self._ext_id_for(p.country_id) else ""),
                        (self._xf_ref("state_id", self._ext_id_for(p.state_id))
                         if p.state_id and self._ext_id_for(p.state_id) else ""),
                        self._xf("vat", p.vat) if p.vat else "",
                    ]))
                    if f_xml.strip():
                        lines.append(self._xml_record(xml_id, "res.partner", f_xml))
        except Exception:
            pass
        if not lines:
            return None
        return self._xml_wrap("\n".join(lines))

    # ── Products XML ────────────────────────────────────────────────────────
    def _modgen_products(self):
        lines = []
        try:
            products = self.env["product.template"].search(
                [("type", "in", ["product", "consu"])],
                limit=500, order="name"
            )
            if products:
                lines.append("    <!-- ── Products: Storable / Consumable ── -->")
                for i, p in enumerate(products):
                    xml_id = f"product_template_{i:04d}"
                    categ_ref = self._ext_id_for(p.categ_id) if p.categ_id else None
                    uom_ref   = self._ext_id_for(p.uom_id) if p.uom_id else None
                    uom_po_ref = self._ext_id_for(p.uom_po_id) if p.uom_po_id else None
                    f_xml  = "\n".join(filter(None, [
                        self._xf("name", p.name) if p.name else "",
                        self._xf("type", p.type),
                        self._xf("sale_ok", "1" if p.sale_ok else "0"),
                        self._xf("purchase_ok", "1" if p.purchase_ok else "0"),
                        self._xf("list_price", p.list_price),
                        self._xf("standard_price", p.standard_price),
                        self._xf_ref("categ_id", categ_ref) if categ_ref else "",
                        self._xf("barcode", p.barcode) if p.barcode else "",
                        self._xf("weight", p.weight) if p.weight else "",
                        self._xf("volume", p.volume) if p.volume else "",
                        self._xf("description_sale", (p.description_sale or "")[:500]) if p.description_sale else "",
                        self._xf_ref("uom_id", uom_ref) if uom_ref else "",
                        self._xf_ref("uom_po_id", uom_po_ref) if uom_po_ref else "",
                    ]))
                    if f_xml.strip():
                        lines.append(self._xml_record(xml_id, "product.template", f_xml))
        except Exception:
            pass
        try:
            # Product attributes
            attrs = self.env["product.attribute"].search([])
            if attrs:
                lines.append("    <!-- ── Product Attributes ── -->")
                for i, a in enumerate(attrs):
                    xml_id = f"product_attr_{re.sub(r'[^a-z0-9]', '_', a.name.lower())}_{i}"
                    f_xml  = "\n".join([
                        self._xf("name", a.name),
                        self._xf("sequence", a.sequence),
                    ])
                    lines.append(self._xml_record(xml_id, "product.attribute", f_xml))
        except Exception:
            pass
        if not lines:
            return None
        return self._xml_wrap("\n".join(lines))

    def _build_sheet(self, wb, area, rows):
        ws = wb.create_sheet(title=area.name[:31])
        ws.sheet_properties.tabColor = (area.tab_color or C["hdr_bg"]).replace("#", "")
        ws.freeze_panes = "A3"
        for i, w in enumerate(COL_W, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.merge_cells("A1:J1")
        ws.row_dimensions[1].height = 30
        c = ws["A1"]
        company_name = self.company_id.name or 'AL3'
        date_str = datetime.now().strftime("%d/%m/%Y")
        c.value = f"{company_name} \u2013 Odoo Config Export | {area.name} | {date_str}"
        c.font      = Font(bold=True, size=13, color=C["hdr_fg"], name="Calibri")
        c.fill      = _fill(C["hdr_bg"])
        c.alignment = _align("center")
        ws.row_dimensions[2].height = 22
        for col, hdr in enumerate(HEADERS, 1):
            cell = ws.cell(row=2, column=col, value=hdr)
            cell.font      = Font(bold=True, color=C["hdr_fg"], size=10, name="Calibri")
            cell.fill      = _fill(C["sub_bg"])
            cell.alignment = _align("center")
            cell.border    = _border()
        prev_area_val = None
        for r_idx, row in enumerate(rows, 3):
            ws.row_dimensions[r_idx].height = 40
            area_val = row[0] if row else ""
            if area_val != prev_area_val:
                row_fill = _fill(C["sect"])
            else:
                row_fill = _fill(C["alt"] if r_idx % 2 == 0 else C["white"])
            prev_area_val = area_val
            for c_idx, val in enumerate(row[:10], 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=str(val) if val is not None else "")
                cell.border    = _border()
                cell.alignment = _align()
                cell.font      = _font(size=9)
                cell.fill      = row_fill
            status = row[8] if len(row) > 8 else "Not Started"
            sc = ws.cell(row=r_idx, column=9)
            bg, fg = _status_fill(status)
            sc.fill      = _fill(bg)
            sc.font      = Font(bold=True, size=9, color=fg, name="Calibri")
            sc.alignment = _align("center")
        ws.add_data_validation(_dv(STATUS_DV, "I"))
        ws.add_data_validation(_dv(MIGR_DV,   "G"))
        ws.add_data_validation(_dv(OWNER_DV,  "H"))
        ws.auto_filter.ref = f"A2:J{len(rows) + 2}"
    # ── Cover sheet ───────────────────────────────────────────────────────────
    def _build_cover(self, wb):
        ws = wb.create_sheet(title="COVER", index=0)
        ws.sheet_properties.tabColor = C["sa_grn"]
        for col, width in zip("ABCDE", [8, 38, 22, 22, 22]):
            ws.column_dimensions[col].width = width
        ws.merge_cells("B1:E1")
        ws.row_dimensions[1].height = 50
        c = ws["B1"]
        company_name = self.company_id.name or 'AL3'
        c.value = f"{company_name} \u2013 Odoo Configuration Export"
        c.font      = Font(bold=True, size=20, color=C["hdr_fg"], name="Calibri")
        c.fill      = _fill(C["hdr_bg"])
        c.alignment = _align("center")
        ws.merge_cells("B2:E2")
        c2 = ws["B2"]
        gen_dt = datetime.now().strftime("%d/%m/%Y %H:%M")
        c2.value = f"Generated: {gen_dt} | {company_name}"
        c2.font      = Font(italic=True, size=11, color=C["hdr_fg"], name="Calibri")
        c2.fill      = _fill(C["sa_grn"])
        c2.alignment = _align("center")
        comp = self.company_id
        meta = [
            ("Company:",   comp.name or ""),
            ("VAT No.:",   comp.vat or "Not Set"),
            ("Country:",   comp.country_id.name if comp.country_id else "Not Set"),
            ("Currency:",  f"{comp.currency_id.name}" if comp.currency_id else "Not Set"),
            ("Phone:",     comp.phone or "Not Set"),
            ("Email:",     comp.email or "Not Set"),
            ("Website:",   comp.website or "Not Set"),
            ("Address:",   ", ".join(filter(None, [
                comp.street,
                comp.city,
                comp.state_id.name if comp.state_id else "",
                comp.zip,
            ])) or "Not Set"),
        ]
        installed = self.env["ir.module.module"].search(
            [("state", "=", "installed"), ("application", "=", True)]
        )
        meta.append(("Installed Apps:", f"{len(installed)} apps installed"))
        meta.append(("Export Date:",    datetime.now().strftime("%d/%m/%Y %H:%M")))
        meta.append(("Exported Areas:", str(len(self.area_ids))))
        for i, (k, v) in enumerate(meta, 4):
            ws.row_dimensions[i].height = 18
            ws[f"B{i}"].value = k
            ws[f"B{i}"].font  = Font(bold=True, size=10, name="Calibri")
            ws[f"C{i}"].value = v
            ws[f"C{i}"].font  = Font(size=10, name="Calibri")
            ws.merge_cells(f"C{i}:E{i}")
        hdr_row = len(meta) + 6
        ws.row_dimensions[hdr_row].height = 22
        for col, txt in enumerate(["#", "Area", "Items (rows)", "Color", "Description"], 2):
            c = ws.cell(row=hdr_row, column=col, value=txt)
            c.font      = Font(bold=True, color=C["hdr_fg"], size=10, name="Calibri")
            c.fill      = _fill(C["sub_bg"])
            c.alignment = _align("center")
            c.border    = _border()
        for i, area in enumerate(self.area_ids.sorted("sequence")):
            r = hdr_row + 1 + i
            ws.row_dimensions[r].height = 18
            ws.cell(row=r, column=2, value=i + 1).font           = _font(bold=True, size=9)
            ws.cell(row=r, column=3, value=area.name).font       = _font(size=9)
            ws.cell(row=r, column=4).fill = _fill(area.tab_color or C["hdr_bg"])
            ws.cell(row=r, column=6, value=area.description or "").font = Font(
                size=9, italic=True, name="Calibri"
            )
            ws.merge_cells(f"F{r}:J{r}")
            for col in range(2, 7):
                ws.cell(row=r, column=col).border    = _border()
                ws.cell(row=r, column=col).alignment = _align()
    # ═════════════════════════════════════════════════════════════════════════
    # DISPATCHER
    # ═════════════════════════════════════════════════════════════════════════
    # ═════════════════════════════════════════════════════════════════════════
    # MENU-DRIVEN EXTRACTOR
    # Content is read live from Odoo. For each area (= root app) we walk its
    # "Configuration" menu branch, and for every menu that opens a window
    # action we dump the records of the model it points at. The menu path is
    # taken straight from Odoo, so nothing is hard-coded per module.
    # ═════════════════════════════════════════════════════════════════════════

    # Models that are transactional or never useful as "configuration" — skipped
    # so we don't dump orders/leads/etc. that happen to hang off a Config menu.
    _SKIP_MODELS = {
        "res.config.settings", "sale.order", "purchase.order", "account.move",
        "crm.lead", "stock.picking", "mrp.production", "project.task",
        "hr.applicant", "hr.expense", "sale.report", "account.report",
        "payment.transaction", "payment.token",
    }
    _RECORD_CAP = 100  # max records listed per config model

    def _root_menu_of_area(self, area):
        if not (area.code or "").startswith("menu_"):
            return self.env["ir.ui.menu"]
        try:
            return self.env["ir.ui.menu"].browse(int(area.code[5:])).exists()
        except ValueError:
            return self.env["ir.ui.menu"]

    def _module_of_menu(self, menu):
        """Owning module name of a menu, via ir.model.data."""
        imd = self.env["ir.model.data"].sudo().search(
            [("model", "=", "ir.ui.menu"), ("res_id", "=", menu.id)], limit=1
        )
        return imd.module or ""

    def _iter_config_leaves(self, menu, path):
        """Yield (full_path, leaf_name, res_model) for window-action leaves."""
        full = f"{path} > {menu.name}" if path else menu.name
        action = menu.action
        if action and action._name == "ir.actions.act_window" and action.res_model:
            yield full, menu.name, action.res_model
        for child in menu.child_id.sorted("sequence"):
            yield from self._iter_config_leaves(child, full)

    def _company_domain(self, model):
        """Restrict to current company when the model is company-aware."""
        Model = self.env[model]
        if "company_id" in Model._fields:
            return ["|", ("company_id", "=", self.company_id.id),
                    ("company_id", "=", False)]
        return []

    def _extract_area(self, area):
        """Build rows for one root app, purely from its menu tree + DB."""
        R = self._row
        root = self._root_menu_of_area(area)
        if not root:
            return []
        rows = []

        # ── 1. Master / config records reachable from the Configuration branch ──
        cfg_branch = self._menu_config_branch(root)
        seen_models = set()
        if cfg_branch:
            base_path = root.name
            for full_path, leaf_name, model in self._iter_config_leaves(
                cfg_branch, base_path
            ):
                if model in self._SKIP_MODELS or model in seen_models:
                    continue
                if model not in self.env:
                    continue
                seen_models.add(model)
                try:
                    Model = self.env[model].sudo()
                    domain = self._company_domain(model)
                    total = Model.search_count(domain)
                    recs = Model.search(domain, limit=self._RECORD_CAP)
                except Exception as e:
                    rows.append(R(area.name, leaf_name, full_path,
                                  "", f"Read error: {e}", "",
                                  status="Blocked", notes=model))
                    continue
                # Summary row for the config type
                rows.append(R(
                    area.name, leaf_name, full_path,
                    "Review / configure",
                    f"{total} record(s)", "",
                    status="Done" if total else "Not Started",
                    notes=model,
                ))
                # One row per record (capped)
                for rec in recs:
                    rows.append(R(
                        area.name, f"• {rec.display_name}", full_path,
                        "", rec.display_name, "",
                        status="Done", notes=model,
                    ))
                if total > len(recs):
                    rows.append(R(
                        area.name, f"… (+{total - len(recs)} more)",
                        full_path, "", "", "",
                        status=None, notes=model,
                    ))

        # ── 2. Live res.config.settings values owned by this app's module ──────
        rows.extend(self._extract_settings_for_area(area, root))
        return rows

    def _extract_settings_for_area(self, area, root):
        """Read res.config.settings fields contributed by this app's module."""
        R = self._row
        module = self._module_of_menu(root)
        if not module:
            return []
        rows = []
        try:
            settings_model = self.env["ir.model"].search(
                [("model", "=", "res.config.settings")], limit=1
            )
            if not settings_model:
                return []
            field_recs = self.env["ir.model.fields"].search([
                ("model_id", "=", settings_model.id),
            ])
            field_names = sorted(
                f.name for f in field_recs
                if (f.modules or "").split(",")[0].strip() == module
                and not f.name.startswith("_")
                and f.name not in ("id", "display_name")
            )
            if not field_names:
                return []
            cfg = self.env["res.config.settings"].sudo().create(
                {"company_id": self.company_id.id}
            )
            defs = cfg.fields_get(field_names)
            settings_path = f"{root.name} > Configuration > Settings"
            for fname in field_names:
                meta = defs.get(fname, {})
                label = meta.get("string") or fname
                ftype = meta.get("type", "")
                try:
                    raw = cfg[fname]
                    if ftype == "many2one":
                        val = raw.display_name if raw else "Not Set"
                    elif ftype in ("many2many", "one2many"):
                        val = ", ".join(raw.mapped("display_name")) or "Not Set"
                    elif ftype == "selection":
                        val = dict(meta.get("selection", [])).get(
                            raw, str(raw)) if raw not in (False, None) else "Not Set"
                    elif ftype == "boolean":
                        val = "Enabled" if raw else "Disabled"
                    else:
                        val = str(raw) if raw not in (None, False, "") else "Not Set"
                except Exception as e:
                    val = f"Read error: {e}"
                rows.append(R(
                    area.name, label, settings_path,
                    "Configure setting", val, "",
                    notes=fname,
                ))
        except Exception as e:
            rows.append(R(area.name, f"Settings read error: {e}",
                          "", "", "", "", status="Blocked"))
        return rows

    # ── helpers ───────────────────────────────────────────────────────────────
    @staticmethod
    def _row(area, item, menu, action, db_val, best,
             migr="No", owner="Management", status=None, notes=""):
        if status is None:
            status = (
                "Done"
                if db_val and str(db_val).strip() not in
                   ("Not Set", "Not Configured", "", "False", "None", "0",
                    "Disabled", "Read error")
                else "Not Started"
            )
        return [area, item, menu, action,
                str(db_val) if db_val is not None else "Not Set",
                best, migr, owner, status, notes]

    def _mod_installed(self, *names):
        return bool(self.env["ir.module.module"].search(
            [("name", "in", list(names)), ("state", "=", "installed")]
        ))

    def _get_depends_list(self, modules):
        """Include modules only if installed. Returns sorted list of names."""
        installed = self.env["ir.module.module"].search(
            [("state", "=", "installed")]
        ).mapped("name")
        result = {m for m in modules if m in installed}
        result.discard("base")  # added separately in manifest
        return sorted(result)
